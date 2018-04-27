import csv
from openpyxl import load_workbook
from datetime import datetime

with open("Results.tsv") as tsv:
    print('Both files found')

    wb = load_workbook("Callouts.xlsx")
    currentSheet = wb['Callouts 2018']

    # Find which row to start appending spreadsheet
    alarmColumn = currentSheet['A']

    print('Working out where to start appending spreadsheet')
    startAppending = False
    i = 2    # Start at 2 as first result will always be 'column header'
    while not startAppending:
        if alarmColumn[i].value is None:
            startAppending = True
            startingRowNumber = i + 1
        i += 1

    print('Will append spreadsheet from row #' + str(startingRowNumber))

    # Skips first row of TSV file due to column headers
    iterResults = iter(csv.reader(tsv, dialect="excel-tab"))
    next(iterResults)

    # Extracting data from TSV file
    i = 0
    for row in iterResults:
        ticketID = row[0]

        ticketCreated = datetime.strptime(row[15], '%Y-%m-%d %H:%M:%S')
        dateCreated = ticketCreated.strftime('%d/%m/%Y')
        timeCreated = ticketCreated.strftime('%H:%M:%S')

        if row[20] == '':    # If no Nagios alarm, use ticket's subject - problem when ceph-mon1-5 callout
            alarm = row[2]
        else:
            alarm = row[20]

        # Putting data into spreadsheet
        currentRow = startingRowNumber + i
        currentSheet.cell(row=currentRow, column=1, value=alarm)          # Alarm name
        currentSheet.cell(row=currentRow, column=3, value=dateCreated)    # Date issued
        currentSheet.cell(row=currentRow, column=4, value=timeCreated)    # Time issued
        currentSheet.cell(row=currentRow, column=5, value=ticketID)       # RT query
        i += 1

    wb.save("Callouts.xlsx")
    print('Spreadsheet changes saved')